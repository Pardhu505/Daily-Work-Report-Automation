import time
import re
import pandas as pd
import openpyxl
from pymongo import MongoClient
from datetime import datetime
import smtplib
from email.message import EmailMessage
import requests
import os
from collections import defaultdict
from openpyxl.styles import Alignment

# ================== CONFIG ==================

MONGO_URI = os.getenv("MONGO_URI")
DB_NAME = "showtime_reports"
COLLECTION_NAME = "work_reports"

SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
SENDER_EMAIL = os.getenv("SMTP_EMAIL")
SENDER_PASSWORD = os.getenv("SMTP_PASSWORD")
RECIPIENTS = ["pardhumunna25@gmail.com"]  # You can add more if needed

TEMPLATE_FILE = "Tasks template.xlsx"      # Commit this file to repo root
OUTPUT_FILE = "Daily_Work_Report.xlsx"

HF_MODEL = "facebook/bart-large-cnn"
HF_API_URL = f"https://api-inference.huggingface.co/models/{HF_MODEL}"
HF_API_KEY = os.getenv("HF_API_KEY")
HF_HEADERS = {"Authorization": f"Bearer {HF_API_KEY}"}

# ============================================

def hf_post(text, timeout=120, max_retries=3):
    payload = {"inputs": text, "parameters": {"max_length": 150, "min_length": 30, "do_sample": False}}
    backoff = 2
    for attempt in range(1, max_retries + 1):
        try:
            resp = requests.post(HF_API_URL, headers=HF_HEADERS, json=payload, timeout=timeout)
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.RequestException as e:
            print(f"[WARN] HF request attempt {attempt}/{max_retries} failed: {e}")
            if attempt == max_retries:
                raise
            sleep_for = backoff ** attempt
            print(f"[INFO] Retrying after {sleep_for}s...")
            time.sleep(sleep_for)
    raise RuntimeError("HF post retries exhausted")

def extract_summary_from_response(resp):
    if resp is None:
        return None
    if isinstance(resp, str):
        return resp
    if isinstance(resp, list) and len(resp) > 0:
        first = resp[0]
        if isinstance(first, dict):
            for key in ("summary_text", "generated_text", "text"):
                if key in first and isinstance(first[key], str):
                    return first[key]
        if isinstance(first, str):
            return first
    if isinstance(resp, dict):
        for key in ("summary_text", "generated_text", "text"):
            if key in resp and isinstance(resp[key], str):
                return resp[key]
    return None

def split_to_bullets(text, max_points=5):
    if not text:
        return []
    sentences = re.split(r'(?<=[\.\!\?])\s+', text.strip())
    clean = [s.strip().rstrip('.!?') for s in sentences if len(s.strip()) > 10]
    if len(clean) == 0:
        parts = re.split(r'[\n;]+', text)
        clean = [p.strip() for p in parts if len(p.strip()) > 10]
    return clean[:max_points]

def bulletify_tasks(tasks_list, max_bullets=5):
    dedup = []
    seen = set()
    for t in tasks_list:
        s = str(t).strip()
        if s and s not in seen:
            dedup.append(s)
            seen.add(s)
        if len(dedup) >= max_bullets:
            break
    return dedup

def summarize_team_tasks(tasks_list, max_points=5):
    dedup = []
    seen = set()
    for t in tasks_list:
        s = str(t).strip()
        if s and s not in seen:
            dedup.append(s)
            seen.add(s)
    if not dedup:
        return "- No tasks reported."

    joined_len = sum(len(x) for x in dedup)
    try:
        if len(dedup) <= 6 and joined_len < 1200:
            text = " . ".join(dedup)
            resp = hf_post(text, timeout=120, max_retries=3)
            summary_text = extract_summary_from_response(resp)
            if summary_text:
                points = split_to_bullets(summary_text, max_points)
                if points:
                    return "\n".join([f"- {p}" for p in points])
    except Exception as e:
        print(f"[WARN] HF single-call summarize failed: {e}")

    chunk_size = 10
    chunk_summaries = []
    for i in range(0, len(dedup), chunk_size):
        chunk = dedup[i:i+chunk_size]
        chunk_text = " . ".join(chunk)
        try:
            resp = hf_post(chunk_text, timeout=120, max_retries=2)
            chunk_summary = extract_summary_from_response(resp)
            chunk_summaries.append(chunk_summary or chunk_text[:1000])
        except Exception as e:
            print(f"[WARN] HF chunk summarization failed (chunk {i//chunk_size}): {e}")
            chunk_summaries.append(" . ".join(chunk))

    try:
        combined = " ".join(chunk_summaries)
        resp2 = hf_post(combined, timeout=120, max_retries=2)
        final_summary_text = extract_summary_from_response(resp2)
        if final_summary_text:
            points = split_to_bullets(final_summary_text, max_points)
            if points:
                return "\n".join([f"- {p}" for p in points])
    except Exception as e:
        print(f"[WARN] HF final summarization failed: {e}")

    fallback = bulletify_tasks(dedup, max_bullets=max_points)
    return "\n".join([f"- {p}" for p in fallback])

def generate_excel_by_team(data):
    team_tasks = defaultdict(list)
    for record in data:
        team = record.get("team", "").strip()
        tasks = record.get("tasks", []) or []
        for t in tasks:
            if isinstance(t, dict):
                details = t.get("details", "").strip()
            else:
                details = str(t).strip()
            if details:
                team_tasks[team].append(details)

    wb = openpyxl.load_workbook(TEMPLATE_FILE)
    ws = wb.active

    title_cell = ws["B4"]
    if title_cell.value and "DD|MM|YYYY" in str(title_cell.value):
        title_cell.value = str(title_cell.value).replace("DD|MM|YYYY", datetime.now().strftime("%d-%b-%Y"))

    for row in range(6, ws.max_row + 1):
        team_cell = ws[f"B{row}"]
        team_name = str(team_cell.value).strip() if team_cell.value else ""
        if team_name and team_name in team_tasks:
            tasks_for_team = team_tasks[team_name]
            print(f"[INFO] Summarizing team '{team_name}' with {len(tasks_for_team)} tasks.")
            summary_text = summarize_team_tasks(tasks_for_team, max_points=5)
            cell = ws[f"G{row}"]
            cell.value = summary_text
            cell.alignment = Alignment(wrap_text=True, horizontal="left", vertical="center")
            lines = summary_text.count("\n") + 1
            ws.row_dimensions[row].height = max(20, min(300, lines * 16))

    wb.save(OUTPUT_FILE)
    print(f"[INFO] Excel report generated: {OUTPUT_FILE}")
    return OUTPUT_FILE

def generate_dept_team_summary(data):
    summary = defaultdict(lambda: {"employees": set(), "reported": set()})
    for record in data:
        dept = record.get("department", "Unknown").strip()
        team = record.get("team", "Unknown").strip()
        emp = record.get("employee_name", "").strip()
        if emp:
            summary[(dept, team)]["employees"].add(emp)
            if record.get("tasks") and len(record.get("tasks")) > 0:
                summary[(dept, team)]["reported"].add(emp)
    rows = ""
    for (dept, team), info in sorted(summary.items()):
        reported = len(info["reported"])
        rows += f"""
            <tr>
                <td style="border:1px solid #ccc; padding:6px; text-align:center;">{dept}</td>
                <td style="border:1px solid #ccc; padding:6px; text-align:center;">{team}</td>
                <td style="border:1px solid #ccc; padding:6px; text-align:center;">{reported}</td>
            </tr>
        """
    return f"""
    <table style="border-collapse:collapse; width:90%;">
      <thead>
        <tr style="background:#f2f2f2;">
          <th style="border:1px solid #ccc; padding:6px; text-align:center;">Department</th>
          <th style="border:1px solid #ccc; padding:6px; text-align:center;">Team</th>
          <th style="border:1px solid #ccc; padding:6px; text-align:center;">Tasks Reported</th>
        </tr>
      </thead>
      <tbody>
        {rows}
      </tbody>
    </table>
    """

def send_email(attachment, data):
    html_table = generate_dept_team_summary(data)
    msg = EmailMessage()
    msg['Subject'] = f"📊 Daily Work Report Summary - {datetime.now().strftime('%d %B %Y')}"
    msg['From'] = SENDER_EMAIL
    msg['To'] = ", ".join(RECIPIENTS)

    html_body = f"""
    <html><body>
      <p>Dear Robbin,</p>
      <p>Please find attached the daily summarized work report.</p>
      <p><b>📅 Department & Team-wise Reporting Summary ({datetime.now().strftime('%d-%b-%Y')})</b></p>
      {html_table}
      <p>Regards,<br/>Pardhasaradhi | Data & Tech Lead<br/>STC-AP</p>
    </body></html>
    """
    msg.set_content("This email requires HTML support.")
    msg.add_alternative(html_body, subtype="html")

    with open(attachment, "rb") as f:
        msg.add_attachment(
            f.read(),
            maintype="application",
            subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=os.path.basename(attachment)
        )
    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as s:
            s.starttls()
            s.login(SENDER_EMAIL, SENDER_PASSWORD)
            s.send_message(msg)
        print("[INFO] Email sent.")
    except Exception as e:
        print(f"[ERROR] Email send failed: {e}")

def fetch_data():
    client = MongoClient(MONGO_URI)
    db = client[DB_NAME]
    coll = db[COLLECTION_NAME]
    today = datetime.now().strftime("%Y-%m-%d")
    data = list(coll.find({"date": today}))
    client.close()
    print(f"[INFO] Fetched {len(data)} records for {today}")
    return data

def export_to_csv(data):
    if not data:
        return None
    df = pd.DataFrame(data)
    if "_id" in df.columns:
        df["_id"] = df["_id"].astype(str)
    path = f"mongo_export_{datetime.now().strftime('%Y%m%d')}.csv"
    df.to_csv(path, index=False)
    print(f"[INFO] Exported to {path}")
    return path

def daily_job():
    print(f"[INFO] Starting daily job at {datetime.now()}")
    data = fetch_data()
    export_to_csv(data)
    try:
        xfile = generate_excel_by_team(data)
    except Exception as e:
        print(f"[ERROR] Excel generation failed: {e}")
        return
    send_email(xfile, data)
    print("[DONE] Daily job completed successfully.")

if __name__ == "__main__":
    daily_job()
